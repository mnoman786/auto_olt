import ipaddress
import socket
from concurrent.futures import ThreadPoolExecutor, as_completed
from django.db.models import Count, Q
from rest_framework import status, generics
from rest_framework.decorators import api_view, permission_classes, throttle_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.throttling import UserRateThrottle
from django.shortcuts import get_object_or_404
from .models import OLT, SetupLog, OLTPort, AutoProvisionConfig
from .serializers import OLTSerializer, OLTListSerializer, OLTCreateSerializer, SetupLogSerializer, OLTPortSerializer, AutoProvisionConfigSerializer
from services import provisioning_service


class OLTSetupThrottle(UserRateThrottle):
    scope = 'olt_setup'


class OLTPollThrottle(UserRateThrottle):
    scope = 'olt_poll'


def get_olt_for_user(pk, user):
    if user.is_staff or user.is_superuser:
        return get_object_or_404(OLT, pk=pk)
    return get_object_or_404(OLT, pk=pk, user=user)


class OLTListCreateView(generics.ListCreateAPIView):
    permission_classes = [IsAuthenticated]

    def get_serializer_class(self):
        if self.request.method == 'POST':
            return OLTCreateSerializer
        return OLTListSerializer

    def get_queryset(self):
        user = self.request.user
        qs = OLT.objects.select_related('user')
        if not (user.is_staff or user.is_superuser):
            qs = qs.filter(user=user)
        q = self.request.query_params.get('search', '').strip()
        if q:
            from django.db.models.functions import Cast
            from django.db.models import CharField
            qs = qs.annotate(_ip_text=Cast('ip_address', output_field=CharField()))
            qs = qs.filter(
                Q(name__icontains=q) |
                Q(_ip_text__icontains=q) |
                Q(system_name__icontains=q)
            )
        return qs.annotate(
            _onu_count=Count('onus', distinct=True),
            _registered_onu_count=Count(
                'onus',
                filter=Q(onus__status__in=('registered', 'active')),
                distinct=True,
            ),
            _vlan_count=Count('vlans', distinct=True),
            _discovered_vlan_count=Count(
                'vlans',
                filter=Q(vlans__source='discovered'),
                distinct=True,
            ),
        ).order_by('-created_at')

    def perform_create(self, serializer):
        olt = serializer.save(user=self.request.user)
        # Only auto-start setup if the OLT is reachable right now.
        # VPN OLTs require the customer's WireGuard peer to be configured first
        # (the Setup page will prompt for it and then start setup manually).
        if olt.connection_type == 'direct' or olt.wg_client_public_key:
            provisioning_service.start_olt_setup_async(olt.id)
        return olt

    def create(self, request, *args, **kwargs):
        from apps.plans.limits import check_olt_limit
        allowed, current, limit = check_olt_limit(request.user)
        if not allowed:
            return Response(
                {
                    'detail': f'OLT limit reached. Your plan allows {limit} OLT(s). '
                              f'Upgrade to add more.',
                    'code': 'olt_limit_reached',
                    'current': current,
                    'limit': limit,
                },
                status=status.HTTP_402_PAYMENT_REQUIRED,
            )
        serializer = self.get_serializer(data=request.data, context={'request': request})
        serializer.is_valid(raise_exception=True)
        olt = self.perform_create(serializer)
        if olt.connection_type == 'vpn' and olt.wg_client_public_key:
            from services import wireguard_service
            wireguard_service.add_peer(olt)
        olt = self.get_queryset().get(pk=olt.pk)
        output = OLTListSerializer(olt, context={'request': request})
        headers = self.get_success_headers(output.data)
        return Response(output.data, status=status.HTTP_201_CREATED, headers=headers)


class OLTDetailView(generics.RetrieveUpdateDestroyAPIView):
    permission_classes = [IsAuthenticated]

    def get_serializer_class(self):
        if self.request.method in ('PUT', 'PATCH'):
            return OLTCreateSerializer
        return OLTSerializer

    def get_queryset(self):
        user = self.request.user
        qs = OLT.objects.select_related('user').annotate(
            _onu_count=Count('onus', distinct=True),
            _registered_onu_count=Count(
                'onus',
                filter=Q(onus__status__in=('registered', 'active')),
                distinct=True,
            ),
            _vlan_count=Count('vlans', distinct=True),
            _discovered_vlan_count=Count(
                'vlans',
                filter=Q(vlans__source='discovered'),
                distinct=True,
            ),
        )
        if not (user.is_staff or user.is_superuser):
            qs = qs.filter(user=user)
        return qs

    def perform_destroy(self, instance):
        if instance.connection_type == 'vpn' and instance.wg_client_public_key:
            from services import wireguard_service
            wireguard_service.remove_peer(instance.wg_client_public_key)
        instance.delete()

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        serializer = self.get_serializer(
            instance, data=request.data, partial=partial,
            context={'request': request}
        )
        serializer.is_valid(raise_exception=True)
        old_pubkey = instance.wg_client_public_key
        olt = serializer.save()
        if olt.connection_type == 'vpn':
            from services import wireguard_service
            if old_pubkey and old_pubkey != olt.wg_client_public_key:
                wireguard_service.remove_peer(old_pubkey)
            if olt.wg_client_public_key:
                wireguard_service.add_peer(olt)
        olt = self.get_queryset().get(pk=olt.pk)
        output = OLTSerializer(olt, context={'request': request})
        return Response(output.data)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
@throttle_classes([OLTSetupThrottle])
def trigger_setup(request, pk):
    """Trigger OLT setup workflow (async)."""
    olt = get_olt_for_user(pk, request.user)
    if olt.status == 'configuring':
        return Response({'detail': 'Setup already in progress.'}, status=400)
    if olt.connection_type == 'vpn' and not olt.wg_client_public_key:
        return Response(
            {'detail': 'Configure the WireGuard peer before starting setup.'},
            status=400,
        )
    provisioning_service.start_olt_setup_async(olt.id)
    return Response({'detail': 'Setup started.', 'olt_id': olt.id})


def _validate_olt_target(host: str, port: int):
    """
    SSRF guard: reject IPs/ports that have no business being an OLT target.
    Returns an error string if the target is disallowed, else None.
    """
    # Port must be in a sane range for Telnet (not scanning arbitrary services)
    if not (1 <= port <= 65535):
        return 'Port must be between 1 and 65535.'
    # Reject non-standard ports that are never used for OLT Telnet
    # (allow 23 and high-port alternatives like 2323, 23000-range)
    ALLOWED_PORTS = {23, 2323, 23231}
    if port not in ALLOWED_PORTS and not (1024 <= port <= 65535):
        return f'Port {port} is not a recognised Telnet port for OLT devices.'

    try:
        addr = ipaddress.ip_address(host)
    except ValueError:
        return 'Invalid IP address format.'

    # Block address categories that are never valid OLT destinations
    blocked = [
        (addr.is_loopback,      'Loopback addresses are not allowed.'),
        (addr.is_link_local,    'Link-local addresses are not allowed (includes cloud metadata 169.254.x.x).'),
        (addr.is_multicast,     'Multicast addresses are not allowed.'),
        (addr.is_unspecified,   'Unspecified address (0.0.0.0) is not allowed.'),
        (addr.is_reserved,      'Reserved addresses are not allowed.'),
    ]
    for condition, msg in blocked:
        if condition:
            return msg

    return None


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def test_connection(request):
    """
    Pre-flight Telnet credential check before adding/saving an OLT.
    Attempts a single Telnet login with the provided credentials, then
    disconnects. Does NOT persist anything to the DB.

    Body: { ip_address, telnet_port?, olt_admin_username, olt_admin_password }
    """
    from services import telnet_service
    from django.conf import settings

    host = (request.data.get('ip_address') or '').strip()
    port_raw = request.data.get('telnet_port') or 23
    username = (request.data.get('olt_admin_username') or '').strip()
    password = request.data.get('olt_admin_password') or ''

    if not host:
        return Response({'success': False, 'message': 'IP address is required.'}, status=400)
    if not username:
        return Response({'success': False, 'message': 'Username is required.'}, status=400)
    if not password:
        return Response({'success': False, 'message': 'Password is required.'}, status=400)

    try:
        port = int(port_raw)
    except (TypeError, ValueError):
        return Response({'success': False, 'message': 'Invalid port number.'}, status=400)

    ssrf_error = _validate_olt_target(host, port)
    if ssrf_error:
        return Response({'success': False, 'message': ssrf_error}, status=400)

    success, message, client = telnet_service.telnet_login(
        host=host,
        username=username,
        password=password,
        port=port,
        timeout=getattr(settings, 'TELNET_TEST_TIMEOUT', 10),
    )
    if client is not None:
        client.disconnect()

    return Response({
        'success': bool(success),
        'message': message or ('Connected.' if success else 'Connection failed.'),
    })


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def reset_status(request, pk):
    """
    Recover an OLT that's stuck in 'configuring' (e.g. after a server
    restart killed the background setup thread). Resets status to
    'pending' so the user can click Retry Setup. Only allowed when the
    OLT is currently in 'configuring'.
    """
    olt = get_olt_for_user(pk, request.user)
    if olt.status != 'configuring':
        return Response(
            {'detail': f'OLT is not stuck (current status: {olt.status}).'},
            status=400,
        )
    olt.status = 'pending'
    olt.save(update_fields=['status'])
    return Response({'detail': 'OLT status reset.', 'olt_id': olt.id, 'status': olt.status})


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def setup_logs(request, pk):
    """Get setup logs for an OLT."""
    olt = get_olt_for_user(pk, request.user)
    logs = olt.setup_logs.all().order_by('created_at')[:200]
    serializer = SetupLogSerializer(logs, many=True)
    return Response({
        'olt_id': olt.id,
        'status': olt.status,
        'logs': serializer.data,
    })


@api_view(['POST'])
@permission_classes([IsAuthenticated])
@throttle_classes([OLTPollThrottle])
def poll_olt(request, pk):
    """Trigger SNMP poll for ONU discovery."""
    from tasks import poll_olt_onus_task
    olt = get_olt_for_user(pk, request.user)
    poll_olt_onus_task.delay(olt.id)
    return Response({'detail': 'Poll started.'})


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def olt_stats(request, pk):
    """Get OLT statistics."""
    olt = get_olt_for_user(pk, request.user)
    agg = olt.onus.aggregate(
        total_onus=Count('id'),
        active_onus=Count('id', filter=Q(status='active')),
        offline_onus=Count('id', filter=Q(status='offline')),
        unregistered_onus=Count('id', filter=Q(status='unregistered')),
        registered_onus=Count('id', filter=Q(status__in=('registered', 'active'))),
    )
    return Response({
        'olt_id': olt.id,
        'status': olt.status,
        **agg,
        'last_polled': olt.last_polled,
    })


@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def olt_ports(request, pk):
    """GET: list saved ports. POST: enqueue background SNMP port discovery."""
    olt = get_olt_for_user(pk, request.user)

    if request.method == 'POST':
        from tasks import discover_ports_task
        discover_ports_task.delay(olt.id)
        ports = olt.ports.all()
        return Response({
            'detail': 'Port discovery queued. Refresh in a few seconds.',
            'count': ports.count(),
            'ports': OLTPortSerializer(ports, many=True).data,
        })

    ports = olt.ports.all()
    return Response({'count': ports.count(), 'ports': OLTPortSerializer(ports, many=True).data})


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def test_snmp(request, pk):
    """
    Diagnostic endpoint — runs all connectivity checks in parallel and returns
    detailed results so the user can see exactly what is failing.
    """
    from services import snmp_service
    from services.provisioning_service import _connect_ip

    olt = get_olt_for_user(pk, request.user)
    connect_ip = _connect_ip(olt)

    def check_tcp(host, port, timeout=3):
        try:
            s = socket.create_connection((host, port), timeout=timeout)
            s.close()
            return True, f'TCP {host}:{port} is reachable'
        except Exception:
            return False, f'TCP port {port} is unreachable'

    def check_snmp_read():
        return snmp_service.validate_snmp_connectivity(
            host=connect_ip, community=olt.snmp_read_community, version=olt.snmp_version,
        )

    def check_snmp_write():
        return snmp_service.validate_snmp_write_access(
            host=connect_ip, write_community=olt.snmp_write_community, version=olt.snmp_version,
        )

    # Run all independent checks concurrently
    futures_map = {}
    with ThreadPoolExecutor(max_workers=4) as pool:
        if olt.telnet_enabled:
            futures_map['tcp_reach'] = pool.submit(check_tcp, connect_ip, olt.telnet_port, 3)
            futures_map['tcp_telnet'] = pool.submit(check_tcp, connect_ip, olt.telnet_port, 4)
        futures_map['snmp_read'] = pool.submit(check_snmp_read)
        if olt.snmp_write_community:
            futures_map['snmp_write'] = pool.submit(check_snmp_write)
        results = {k: f.result() for k, f in futures_map.items()}

    checks = []

    if 'tcp_reach' in results:
        ok, detail = results['tcp_reach']
        checks.append({'check': 'network_reach', 'ok': ok, 'detail': detail})

    snmp_result = results.get('snmp_read', {})
    checks.append({
        'check': 'snmp_read',
        'ok': snmp_result.get('connected', False),
        'detail': (
            f'sysDescr = {snmp_result["sys_descr"][:120]}' if snmp_result.get('connected')
            else 'SNMP GET failed — check read community and SNMP version'
        ),
        'oid_tested': '1.3.6.1.2.1.1.1.0 (sysDescr)',
        'snmp_version': olt.snmp_version,
    })

    if 'snmp_write' in results:
        write_result = results['snmp_write']
        checks.append({
            'check': 'snmp_write',
            'ok': write_result.get('writable', False),
            'detail': (
                'Write access confirmed' if write_result.get('writable')
                else 'SNMP write failed — check write community and OLT permissions'
            ),
        })
    else:
        checks.append({'check': 'snmp_write', 'ok': None,
                        'detail': 'No write community configured — skipped'})

    if 'tcp_telnet' in results:
        ok, detail = results['tcp_telnet']
        checks.append({'check': 'telnet_port', 'ok': ok,
                        'detail': detail.replace('reachable', 'open').replace('unreachable', 'closed or refused')})

    overall_ok = all(c['ok'] for c in checks if c['ok'] is not None)
    return Response({
        'olt_id': olt.id,
        'olt_name': olt.name,
        'ip_address': olt.ip_address,
        'overall': 'pass' if overall_ok else 'fail',
        'checks': checks,
        'hint': (
            'All checks passed — OLT is reachable and SNMP is working.' if overall_ok else
            'One or more checks failed. Common causes: '
            '(1) Wrong IP or OLT is offline, '
            '(2) Firewall blocking UDP 161, '
            '(3) Wrong SNMP community string, '
            '(4) SNMP agent not enabled on the OLT device itself.'
        ),
    })


@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def wg_info(request, pk):
    """
    GET: Return WireGuard connection info for this OLT (server pubkey, endpoint, peer status).
    POST: Update client public key + subnet, then sync WireGuard peer.
    """
    from services import wireguard_service
    olt = get_olt_for_user(pk, request.user)

    if olt.connection_type != 'vpn':
        return Response({'detail': 'This OLT is not configured for VPN (WireGuard).'}, status=400)

    if request.method == 'POST':
        import re
        old_pubkey = olt.wg_client_public_key
        new_pubkey = request.data.get('wg_client_public_key', '').strip()
        new_subnet = request.data.get('wg_client_subnet', '').strip()

        if new_pubkey and not re.fullmatch(r'[A-Za-z0-9+/]{43}=', new_pubkey):
            return Response(
                {'error': 'Invalid WireGuard public key format. Must be 44-character base64.'},
                status=400,
            )

        if old_pubkey and old_pubkey != new_pubkey:
            wireguard_service.remove_peer(old_pubkey)

        olt.wg_client_public_key = new_pubkey
        olt.wg_client_subnet = new_subnet
        olt.save(update_fields=['wg_client_public_key', 'wg_client_subnet'])

        if new_pubkey:
            success, msg = wireguard_service.add_peer(olt)
            if not success:
                return Response({'error': msg}, status=400)

    info = wireguard_service.get_wg_info(olt)
    return Response(info)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def bandwidth(request, pk):
    """
    GET /olts/<id>/bandwidth/?hours=24&port_id=<id>
    Returns bandwidth samples grouped by port for the requested window.

    Downsampling — keeps the response small regardless of window size:
      <=6h  : every sample (~5-min resolution, max ~72 pts/port)
      <=24h : every 3rd sample (~15-min resolution, max ~96 pts/port)
      <=48h : every 6th sample (~30-min resolution, max ~96 pts/port)
      >48h  : every 12th sample (~1-hour resolution, max ~168 pts/port)
    """
    from .models import BandwidthSample
    olt = get_olt_for_user(pk, request.user)

    try:
        hours = max(1, min(int(request.query_params.get('hours', 24)), 168))
    except (TypeError, ValueError):
        hours = 24

    from django.utils import timezone
    from datetime import timedelta
    since = timezone.now() - timedelta(hours=hours)

    # Pick downsample stride based on window
    if hours <= 6:
        stride = 1
    elif hours <= 24:
        stride = 3
    elif hours <= 48:
        stride = 6
    else:
        stride = 12

    qs = (
        BandwidthSample.objects
        .filter(port__olt=olt, timestamp__gte=since)
        .select_related('port')
        .order_by('port_id', 'timestamp')
    )

    port_id_filter = request.query_params.get('port_id')
    if port_id_filter:
        try:
            qs = qs.filter(port_id=int(port_id_filter))
        except (TypeError, ValueError):
            pass

    from django.core.cache import cache
    import json

    cache_key = f'bw_api:{olt.id}:h{hours}' + (f':p{port_id_filter}' if port_id_filter else '')
    cached = cache.get(cache_key)
    if cached:
        return Response(cached)

    ports_data: dict = {}
    counters: dict[int, int] = {}

    for sample in qs:
        pid = sample.port_id
        if pid not in ports_data:
            ports_data[pid] = {
                'port_id':   pid,
                'port_name': sample.port.name,
                'port_type': sample.port.port_type,
                'samples':   [],
            }
            counters[pid] = 0

        if counters[pid] % stride == 0:
            ports_data[pid]['samples'].append({
                't':        sample.timestamp.isoformat(),
                'in_mbps':  sample.in_mbps,
                'out_mbps': sample.out_mbps,
            })
        counters[pid] += 1

    payload = {
        'olt_id': olt.id,
        'hours':  hours,
        'ports':  list(ports_data.values()),
    }
    # Cache for 4 min — slightly under the 5-min poll interval so data stays fresh
    cache.set(cache_key, payload, timeout=240)
    return Response(payload)


@api_view(['GET', 'PUT'])
@permission_classes([IsAuthenticated])
def auto_provision(request, pk):
    """
    GET  /olts/<id>/auto-provision/  — return current config
    PUT  /olts/<id>/auto-provision/  — update config
         Body: { enabled, default_vlan (id or null), line_profile_id, srv_profile_id }
    """
    olt = get_olt_for_user(pk, request.user)
    config, _ = AutoProvisionConfig.objects.get_or_create(olt=olt)

    if request.method == 'GET':
        return Response(AutoProvisionConfigSerializer(config).data)

    # Validate that the VLAN belongs to this OLT
    vlan_id = request.data.get('default_vlan')
    if vlan_id:
        from apps.vlans.models import VLAN
        if not VLAN.objects.filter(id=vlan_id, olt=olt).exists():
            return Response({'default_vlan': 'VLAN does not belong to this OLT.'}, status=400)

    serializer = AutoProvisionConfigSerializer(config, data=request.data, partial=True)
    serializer.is_valid(raise_exception=True)
    serializer.save()
    return Response(AutoProvisionConfigSerializer(config).data)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def sync_profiles(request, pk):
    """
    Re-read ONU line + service profiles from the OLT and cache them.
    Used so ONU registration can auto-pick a valid profile ID without
    the user having to know what IDs exist on the device.
    """
    olt = get_olt_for_user(pk, request.user)
    result = provisioning_service.sync_profiles_from_olt(olt.id)
    if not result.get('success'):
        return Response(
            {'detail': result.get('error') or 'Profile sync failed', **result},
            status=400,
        )
    return Response(result)



@api_view(['GET'])
@permission_classes([IsAuthenticated])
def olt_report_excel(request, pk):
    """Download an Excel report for a single OLT: summary + ONU list + port utilization."""
    import io
    from django.http import HttpResponse
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    olt = get_olt_for_user(pk, request.user)
    onus = list(olt.onus.select_related('vlan').order_by('pon_port', 'onu_index'))
    ports = list(olt.ports.order_by('port_type', 'name'))

    wb = openpyxl.Workbook()

    # ── Sheet 1: Summary ──────────────────────────────────────────────────────
    ws = wb.active
    ws.title = 'Summary'
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(fill_type='solid', fgColor='1D4ED8')

    def hdr(ws, row, col, val):
        c = ws.cell(row=row, column=col, value=val)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal='center')

    summary_rows = [
        ('OLT Name', olt.name),
        ('IP Address', olt.ip_address),
        ('Status', olt.status.upper()),
        ('System Name', olt.system_name or 'N/A'),
        ('Uptime', olt.system_uptime or 'N/A'),
        ('Last Polled', olt.last_polled.isoformat() if olt.last_polled else 'Never'),
        ('Total ONUs', len(onus)),
        ('Active ONUs', sum(1 for o in onus if o.status == 'active')),
        ('Offline ONUs', sum(1 for o in onus if o.status == 'offline')),
        ('Unregistered ONUs', sum(1 for o in onus if o.status == 'unregistered')),
        ('PON Ports', sum(1 for p in ports if p.port_type == 'pon')),
        ('Report Generated', __import__('datetime').datetime.now().strftime('%Y-%m-%d %H:%M')),
    ]
    hdr(ws, 1, 1, 'Field')
    hdr(ws, 1, 2, 'Value')
    for i, (k, v) in enumerate(summary_rows, start=2):
        ws.cell(row=i, column=1, value=k).font = Font(bold=True)
        ws.cell(row=i, column=2, value=v)
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 35

    # ── Sheet 2: ONU List ─────────────────────────────────────────────────────
    ws2 = wb.create_sheet('ONU List')
    onu_headers = ['Serial Number', 'MAC Address', 'PON Port', 'ONU Index',
                   'Status', 'Signal (dBm)', 'VLAN', 'Description', 'Last Seen', 'Registered At']
    for col, h in enumerate(onu_headers, 1):
        hdr(ws2, 1, col, h)
    status_colors = {'active': 'D1FAE5', 'offline': 'FEE2E2', 'unregistered': 'FEF3C7',
                     'provisioning': 'DBEAFE', 'registered': 'EDE9FE'}
    for row, o in enumerate(onus, start=2):
        vals = [o.serial_number, o.mac_address, o.pon_port, o.onu_index, o.status,
                o.signal_strength, f'{o.vlan.vlan_id} {o.vlan.name}' if o.vlan else '',
                o.description,
                o.last_seen.isoformat() if o.last_seen else '',
                o.registered_at.isoformat() if o.registered_at else '']
        fill_color = status_colors.get(o.status, 'FFFFFF')
        for col, v in enumerate(vals, 1):
            c = ws2.cell(row=row, column=col, value=v)
            if col == 5:
                c.fill = PatternFill(fill_type='solid', fgColor=fill_color)
    for col in range(1, len(onu_headers) + 1):
        ws2.column_dimensions[get_column_letter(col)].width = 18

    # ── Sheet 3: Port Utilization ─────────────────────────────────────────────
    ws3 = wb.create_sheet('Port Utilization')
    port_headers = ['Port Name', 'Type', 'Status', 'ONUs', 'Max Capacity', 'Utilization %']
    for col, h in enumerate(port_headers, 1):
        hdr(ws3, 1, col, h)
    for row, p in enumerate(ports, start=2):
        util = p.utilization_pct
        vals = [p.name, p.port_type, p.status, p.onu_count, p.max_capacity, util]
        for col, v in enumerate(vals, 1):
            c = ws3.cell(row=row, column=col, value=v)
            if col == 6 and util is not None:
                if util >= 80:
                    c.fill = PatternFill(fill_type='solid', fgColor='FEE2E2')
                elif util >= 60:
                    c.fill = PatternFill(fill_type='solid', fgColor='FEF3C7')
                else:
                    c.fill = PatternFill(fill_type='solid', fgColor='D1FAE5')
    for col in range(1, len(port_headers) + 1):
        ws3.column_dimensions[get_column_letter(col)].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f'OLT_{olt.name}_report.xlsx'.replace(' ', '_')
    response = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{fname}"'
    return response
